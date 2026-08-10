from openpyxl import load_workbook
from dataclasses import dataclass
from io import BytesIO


@dataclass
class BranRow:
    nombre: str
    perfil: str
    servicio_nombre: str


class BranParserError(Exception):
    def __init__(self, message: str, details: list[str] | None = None):
        super().__init__(message)
        self.details = details or []


def parse_bran_excel(file_content: bytes) -> list[BranRow]:
    """Parse BRAN Excel export. Expected columns: Nombre, Perfil, Servicio."""
    try:
        wb = load_workbook(BytesIO(file_content), read_only=True)
    except Exception as e:
        raise BranParserError(f"No se pudo abrir el archivo Excel: {e}")

    ws = wb.active
    if ws is None:
        raise BranParserError("El archivo no contiene hojas de trabajo")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise BranParserError("El archivo está vacío o solo contiene cabecera")

    # Validate header
    header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
    required = {"nombre", "perfil", "servicio"}
    found = set(header)
    missing = required - found
    if missing:
        raise BranParserError(
            f"Columnas requeridas no encontradas: {', '.join(missing)}",
            details=[f"Cabecera encontrada: {rows[0]}"]
        )

    nombre_idx = header.index("nombre")
    perfil_idx = header.index("perfil")
    servicio_idx = header.index("servicio")

    results = []
    errors = []
    for i, row in enumerate(rows[1:], start=2):
        nombre = str(row[nombre_idx]).strip() if row[nombre_idx] else ""
        perfil = str(row[perfil_idx]).strip() if row[perfil_idx] else ""
        servicio = str(row[servicio_idx]).strip() if row[servicio_idx] else ""

        if not nombre or not perfil or not servicio:
            errors.append(f"Fila {i}: campos vacíos (nombre='{nombre}', perfil='{perfil}', servicio='{servicio}')")
            continue

        results.append(BranRow(nombre=nombre, perfil=perfil, servicio_nombre=servicio))

    if errors and not results:
        raise BranParserError("No se encontraron filas válidas", details=errors)

    return results
